import sqlite3
import asyncio
import time
import os
import csv
import openpyxl
from openpyxl.styles import Font, Alignment
from openpyxl.worksheet.filters import AutoFilter
from datetime import datetime, timedelta
from telegram import Update, ReplyKeyboardMarkup, ReplyKeyboardRemove
from telegram.ext import (
    Application, CommandHandler, MessageHandler, filters, ContextTypes, ConversationHandler
)
from telegram.error import TimedOut


# Substitua pelo seu token
TOKEN = '7984935357:AAF1i8h6Q3nBaA97tsd1tsRX-LtpxKkKEhw'

# Caminho do banco de dados no OneDrive
DB_PATH = 'study_tracker.db'

# Aguarde a sincronização do OneDrive (opcional, ajuste conforme necessário)
print("Aguardando sincronização do OneDrive por 5 segundos...")
time.sleep(5)  # Aguarda 5 segundos para garantir que o OneDrive sincronize o arquivo

# Conectar ao banco de dados SQLite
try:
    conn = sqlite3.connect(DB_PATH, check_same_thread=False)
    cursor = conn.cursor()
    print(f"Banco de dados {DB_PATH} aberto com sucesso!")
except sqlite3.Error as e:
    print(f"Erro ao abrir o banco de dados {DB_PATH}: {e}")
    raise

# Criar tabelas se não existirem
cursor.execute('''
    CREATE TABLE IF NOT EXISTS studies (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        user_id INTEGER,
        subject TEXT,
        topic TEXT,
        questions_studied INTEGER,
        study_date TEXT
    )
''')

cursor.execute('''
    CREATE TABLE IF NOT EXISTS reviews (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        study_id INTEGER,
        review_number INTEGER,
        review_date TEXT,
        review_questions INTEGER,
        FOREIGN KEY (study_id) REFERENCES studies(id)
    )
''')

# Adicionar colunas faltantes à tabela studies
try:
    cursor.execute('ALTER TABLE studies ADD COLUMN category TEXT')
except sqlite3.OperationalError:
    pass  # Coluna já existe

try:
    cursor.execute('ALTER TABLE studies ADD COLUMN notes TEXT')
except sqlite3.OperationalError:
    pass  # Coluna já existe

# Adicionar colunas faltantes à tabela reviews
try:
    cursor.execute('ALTER TABLE reviews ADD COLUMN completed INTEGER DEFAULT 0')
except sqlite3.OperationalError:
    pass  # Coluna já existe

try:
    cursor.execute('ALTER TABLE reviews ADD COLUMN priority TEXT DEFAULT "média"')
except sqlite3.OperationalError:
    pass  # Coluna já existe

try:
    cursor.execute('ALTER TABLE reviews ADD COLUMN notes TEXT')
except sqlite3.OperationalError:
    pass  # Coluna já existe

conn.commit()

# Função para converter data de DD-MM-YYYY para YYYY-MM-DD
def convert_to_db_format(date_str):
    try:
        date_obj = datetime.strptime(date_str, '%d-%m-%Y')
        return date_obj.strftime('%Y-%m-%d')
    except ValueError as e:
        raise ValueError(f"Erro ao converter data: {e}")

# Função para converter data de YYYY-MM-DD para DD-MM-YYYY
def convert_to_display_format(date_str):
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.strftime('%d-%m-%Y')
    except ValueError as e:
        raise ValueError(f"Erro ao converter data: {e}")

# Função para calcular datas de revisão
def calculate_review_date(study_date, days):
    try:
        # Converter a data de estudo para o formato interno (YYYY-MM-DD)
        study_date = convert_to_db_format(study_date) if '-' in study_date and study_date[2] == '-' else study_date
        study_date = datetime.strptime(study_date, '%Y-%m-%d')
        review_date = study_date + timedelta(days=days)
        # Retornar no formato interno (YYYY-MM-DD) para armazenamento
        return review_date.strftime('%Y-%m-%d')
    except ValueError as e:
        raise ValueError(f"Erro ao calcular data de revisão: {e}")

# Função para calcular o intervalo de dias entre duas datas
def calculate_days_difference(start_date, end_date):
    try:
        # Converter as datas para o formato interno (YYYY-MM-DD) se necessário
        start_date = convert_to_db_format(start_date) if '-' in start_date and start_date[2] == '-' else start_date
        end_date = convert_to_db_format(end_date) if '-' in end_date and end_date[2] == '-' else end_date
        start = datetime.strptime(start_date, '%Y-%m-%d')
        end = datetime.strptime(end_date, '%Y-%m-%d')
        return (end - start).days
    except ValueError as e:
        raise ValueError(f"Erro ao calcular diferença de dias: {e}")

from telegram import InlineKeyboardButton, InlineKeyboardMarkup


# Estados da conversa para /add
SUBJECT, TOPIC, CATEGORY, QUESTIONS_YN, QUESTIONS_AMOUNT, NUM_REVIEWS, REVIEW_DAYS, REVIEW_QUESTIONS, REVIEW_PRIORITY, STUDY_NOTES, REVIEW_NOTES = range(11)

# Estados da conversa para /edit
EDIT_STUDY_ID, EDIT_DATE_YN, EDIT_DATE, EDIT_NUM_REVIEWS, EDIT_REVIEW_DAYS, EDIT_REVIEW_QUESTIONS, EDIT_REVIEW_PRIORITY = range(7)

# Estados da conversa para /redo
REDO_STUDY_ID, REDO_NUM_REVIEWS, REDO_REVIEW_DAYS, REDO_REVIEW_QUESTIONS, REDO_REVIEW_PRIORITY = range(5)

# Estados da conversa para /delete
DELETE_STUDY_ID, DELETE_CONFIRM = range(2)

# Estados da conversa para /markdone
MARKDONE_REVIEW_ID, MARKDONE_CONFIRM = range(2)

# Comando /start
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "*Oi! Eu sou seu StudyBot.* 😊 Me diga o que você estudou hoje com o comando /add, "
        "e eu te ajudo a planejar suas revisões. Use /check para ver suas revisões pendentes! "
        "Use /summary para ver um resumo das disciplinas estudadas e revisões. "
        "Use /edit para editar os dias e questões das revisões. "
        "Use /redo para refazer o plano de revisões. "
        "Use /delete para excluir um estudo. "
        "Use /stats para ver estatísticas de estudo. "
        "Use /markdone para marcar revisões como concluídas. "
        "Use /export para exportar seus dados para CSV. "
        "Use /weeklyreport para ver um relatório semanal. "
        "Use /start para voltar ao início. "
        "Use /stop para parar o bot. 📅",
        parse_mode='Markdown',
        reply_markup=ReplyKeyboardRemove()  # Remove qualquer teclado anterior
    )

async def stop(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.effective_message.reply_text("Bot parado. Até logo! 👋", reply_markup=ReplyKeyboardRemove())
    await context.application.stop()
    await context.application.shutdown()

# Início da conversa com /add
async def add_study(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['study'] = {}  # Inicializa um dicionário para armazenar as respostas
    context.user_data['reviews'] = []  # Lista para armazenar as revisões
    context.user_data['current_review'] = 0  # Contador para as revisões
    message = update.effective_message  # Usa effective_message para compatibilidade com CallbackQuery
    await message.reply_text("Qual matéria você estudou hoje? (Ex.: Matemática)", reply_markup=ReplyKeyboardRemove())
    return SUBJECT

# Etapa 1: Receber a matéria
async def get_subject(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['study']['subject'] = update.message.text
    await update.message.reply_text("Qual foi o assunto que você estudou? (Ex.: Álgebra)")
    return TOPIC

# Etapa 2: Receber o assunto
async def get_topic(update: Update, context: ContextTypes.DEFAULT_TYPE):
    context.user_data['study']['topic'] = update.message.text
    await update.message.reply_text("Qual a categoria deste estudo? (Ex.: Concursos, Faculdade, Pessoal. Digite 'Nenhuma' para pular.)")
    return CATEGORY

# Etapa 3: Receber a categoria
async def get_category(update: Update, context: ContextTypes.DEFAULT_TYPE):
    category = update.message.text
    if category.lower() == 'nenhuma':
        context.user_data['study']['category'] = None
    else:
        context.user_data['study']['category'] = category
    await update.message.reply_text("Você fez alguma questão sobre esse assunto? Responda 'sim' ou 'não'.")
    return QUESTIONS_YN

# Etapa 4: Perguntar se fez questões
async def get_questions_yn(update: Update, context: ContextTypes.DEFAULT_TYPE):
    response = update.message.text.lower()
    if response not in ['sim', 'não', 'nao']:
        await update.message.reply_text("Por favor, responda 'sim' ou 'não'.")
        return QUESTIONS_YN
    context.user_data['study']['questions_yn'] = response
    if response == 'sim':
        await update.message.reply_text("Quantas questões você fez? (Digite um número, ex.: 20)")
        return QUESTIONS_AMOUNT
    else:
        context.user_data['study']['questions_studied'] = 0
        await update.message.reply_text("Quantas revisões você deseja programar? (Ex.: 3)")
        return NUM_REVIEWS

# Etapa 5: Receber a quantidade de questões (se sim)
async def get_questions_amount(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        questions = int(update.message.text)
        if questions < 0:
            raise ValueError
        context.user_data['study']['questions_studied'] = questions
        await update.message.reply_text("Quantas revisões você deseja programar? (Ex.: 3)")
        return NUM_REVIEWS
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido (ex.: 20).")
        return QUESTIONS_AMOUNT

# Etapa 6: Receber o número de revisões
async def get_num_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        num_reviews = int(update.message.text)
        if num_reviews < 1:
            raise ValueError
        context.user_data['study']['num_reviews'] = num_reviews
        context.user_data['current_review'] = 1
        await update.message.reply_text(f"Em quantos dias você quer fazer a {context.user_data['current_review']}ª revisão? (Ex.: 3)")
        return REVIEW_DAYS
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido maior que 0 (ex.: 3).")
        return NUM_REVIEWS

# Etapa 7: Receber os dias para a revisão atual
async def get_review_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        days = int(update.message.text)
        if days < 0:
            raise ValueError
        context.user_data['reviews'].append({'days': days})
        await update.message.reply_text(f"Quantas questões você quer responder na {context.user_data['current_review']}ª revisão? (Ex.: 15)")
        return REVIEW_QUESTIONS
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido (ex.: 3).")
        return REVIEW_DAYS

# Etapa 8: Receber as questões para a revisão atual
async def get_review_questions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        questions = int(update.message.text)
        if questions < 0:
            raise ValueError
        context.user_data['reviews'][-1]['questions'] = questions
        await update.message.reply_text(f"Qual a prioridade desta revisão? (alta, média, baixa. Digite 'média' para padrão.)")
        return REVIEW_PRIORITY
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido (ex.: 15).")
        return REVIEW_QUESTIONS

# Etapa 9: Receber a prioridade da revisão
async def get_review_priority(update: Update, context: ContextTypes.DEFAULT_TYPE):
    priority = update.message.text.lower()
    if priority not in ['alta', 'média', 'baixa']:
        await update.message.reply_text("Por favor, escolha uma prioridade válida: alta, média ou baixa.")
        return REVIEW_PRIORITY
    context.user_data['reviews'][-1]['priority'] = priority
    await update.message.reply_text(f"Alguma nota para esta revisão? (Digite a nota ou 'Nenhuma' para pular.)")
    return REVIEW_NOTES

# Etapa 10: Receber notas para a revisão
async def get_review_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    notes = update.message.text
    if notes.lower() == 'nenhuma':
        context.user_data['reviews'][-1]['notes'] = None
    else:
        context.user_data['reviews'][-1]['notes'] = notes

    # Incrementar o contador de revisões
    context.user_data['current_review'] += 1

    # Verificar se ainda há revisões a serem adicionadas
    if context.user_data['current_review'] <= context.user_data['study']['num_reviews']:
        await update.message.reply_text(f"Em quantos dias você quer fazer a {context.user_data['current_review']}ª revisão? (Ex.: 3)")
        return REVIEW_DAYS
    else:
        await update.message.reply_text("Alguma nota para este estudo? (Digite a nota ou 'Nenhuma' para pular.)")
        return STUDY_NOTES

# Etapa 11: Receber notas para o estudo e finalizar
async def get_study_notes(update: Update, context: ContextTypes.DEFAULT_TYPE):
    notes = update.message.text
    if notes.lower() == 'nenhuma':
        context.user_data['study']['notes'] = None
    else:
        context.user_data['study']['notes'] = notes

    # Salvar o estudo no banco de dados
    user_id = update.message.from_user.id
    study = context.user_data['study']
    study_date = datetime.now().strftime('%Y-%m-%d')
    study_date_display = convert_to_display_format(study_date)

    cursor.execute('''
        INSERT INTO studies (user_id, subject, topic, questions_studied, study_date, category, notes)
        VALUES (?, ?, ?, ?, ?, ?, ?)
    ''', (user_id, study['subject'], study['topic'], study['questions_studied'], study_date, study['category'], study['notes']))
    study_id = cursor.lastrowid

    # Salvar as revisões
    message = "*Estudo registrado!* Você estudou *" + study['topic'] + "* (*" + study['subject'] + "*) e fez " + str(study['questions_studied']) + " questões."
    if study['category']:
        message += f"\nCategoria: {study['category']}"
    if study['notes']:
        message += f"\nNotas: {study['notes']}"
    message += "\nSuas revisões estão marcadas para:\n"
    for i, review in enumerate(context.user_data['reviews'], 1):
        review_date = calculate_review_date(study_date, review['days'])
        review_date_display = convert_to_display_format(review_date)
        cursor.execute('''
            INSERT INTO reviews (study_id, review_number, review_date, review_questions, priority, notes)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (study_id, i, review_date, review['questions'], review['priority'], review['notes']))
        message += f"Revisão {i}: {review_date_display} ({review['questions']} questões, Prioridade: {review['priority']})"
        if review['notes']:
            message += f"\n  Notas: {review['notes']}"
        message += "\n"

    conn.commit()

    await update.message.reply_text(message, parse_mode='Markdown')

    # Limpar os dados temporários
    context.user_data.clear()
    return ConversationHandler.END

# Cancelar a conversa
async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text("Registro cancelado. Use /add para começar de novo!")
    context.user_data.clear()
    return ConversationHandler.END

# Comando /check para verificar revisões pendentes
async def check_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id  # Usa effective_user para compatibilidade
    today = datetime.now().strftime('%Y-%m-%d')
    today_display = convert_to_display_format(today)

    cursor.execute('''
        SELECT s.subject, s.topic, r.review_date, r.review_questions, r.review_number, r.priority, r.notes
        FROM studies s
        JOIN reviews r ON s.id = r.study_id
        WHERE s.user_id = ? AND r.review_date = ? AND r.completed = 0
    ''', (user_id, today))
    reviews = cursor.fetchall()

    message = f"*📅 Revisões do Dia (Hoje: {today_display}):*\n\n"
    message += "Use este comando para verificar suas revisões pendentes diariamente.\n\n"
    if not reviews:
        await update.effective_message.reply_text("Nenhuma revisão pendente para hoje! 😊", parse_mode='Markdown')
        return

    review_entries = []
    for review in reviews:
        subject, topic, review_date, review_questions, review_number, priority, notes = review
        review_date_display = convert_to_display_format(review_date)
        entry = f"- *{subject} ({topic})*: Revisão {review_number} hoje! Responda {review_questions} questões."
        if priority == 'alta':
            entry += " ⚠️ (Alta prioridade)"
        if notes:
            entry += f"\n  Notas: {notes}"
        entry += "\n"
        review_entries.append(entry)

    # Dividir as revisões em mensagens menores, se necessário
    MAX_MESSAGE_LENGTH = 4096  # Limite do Telegram
    current_message = message
    messages = []
    for entry in review_entries:
        if len(current_message) + len(entry) <= MAX_MESSAGE_LENGTH:
            current_message += entry
        else:
            messages.append(current_message)
            current_message = message + entry
    if current_message != message:
        messages.append(current_message)

    # Enviar as mensagens
    for msg in messages:
        await update.effective_message.reply_text(msg, parse_mode='Markdown')
        await asyncio.sleep(0.5)  # Pequeno atraso para evitar flood

async def summary(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    today = datetime.now().strftime('%Y-%m-%d')
    today_display = convert_to_display_format(today)

    # Verificar se há uma categoria especificada
    args = context.args
    category_filter = args[0] if args else None

    # Dicionário para agrupar informações por data
    events_by_date = {}

    # Parte 1: Aulas estudadas
    query = '''
        SELECT subject, topic, study_date, category, notes
        FROM studies
        WHERE user_id = ?
    '''
    params = [user_id]
    if category_filter:
        query += " AND category = ?"
        params.append(category_filter)

    cursor.execute(query, params)
    studies = cursor.fetchall()

    for study in studies:
        subject, topic, study_date, category, notes = study
        study_date_display = convert_to_display_format(study_date)
        if study_date not in events_by_date:
            events_by_date[study_date] = {'studies': [], 'reviews': []}
        study_entry = f"- *{subject} ({topic})*"
        if category:
            study_entry += f" (Categoria: {category})"
        if notes:
            study_entry += f"\n  Notas: {notes}"
        events_by_date[study_date]['studies'].append(study_entry)

    # Parte 2: Revisões (passadas, atuais e futuras)
    query = '''
        SELECT s.subject, s.topic, r.review_date, r.review_questions, r.review_number, r.priority, r.notes, r.completed
        FROM studies s
        JOIN reviews r ON s.id = r.study_id
        WHERE s.user_id = ?
    '''
    params = [user_id]
    if category_filter:
        query += " AND s.category = ?"
        params.append(category_filter)

    cursor.execute(query, params)
    reviews = cursor.fetchall()

    for review in reviews:
        subject, topic, review_date, review_questions, review_number, priority, notes, completed = review
        review_date_display = convert_to_display_format(review_date)
        if review_date not in events_by_date:
            events_by_date[review_date] = {'studies': [], 'reviews': []}
        review_entry = f"- *{subject} ({topic})*: Revisão {review_number} - Responda {review_questions} questões"
        if priority == 'alta':
            review_entry += " ⚠️ (Alta prioridade)"
        if completed:
            review_entry += " ✅ (Concluída)"
        else:
            if review_date == today:
                review_entry += " (Hoje!)"
            elif review_date < today:
                review_entry += " (Atrasada)"
        if notes:
            review_entry += f"\n  Notas: {notes}"
        events_by_date[review_date]['reviews'].append(review_entry)

    # Ordenar as datas
    sorted_dates = sorted(events_by_date.keys())

    # Enviar mensagens separadas por data
    MAX_MESSAGE_LENGTH = 4096  # Limite do Telegram
    for date in sorted_dates:
        date_display = convert_to_display_format(date)
        message = f"*📅 Data: {date_display}*\n\n"

        # Aulas estudadas
        study_entries = events_by_date[date]['studies']
        if study_entries:
            message += "*Aulas Estudadas:*\n"
            study_message = "\n".join(study_entries) + "\n"
        else:
            study_message = "*Aulas Estudadas:* Nenhuma aula estudada nesta data.\n"

        # Revisões
        review_entries = events_by_date[date]['reviews']
        if review_entries:
            review_message = "\n*Revisões Programadas:*\n" + "\n".join(review_entries) + "\n"
        else:
            review_message = "\n*Revisões Programadas:* Nenhuma revisão programada para esta data.\n"

        # Combinar as partes e dividir se necessário
        full_message = message + study_message + review_message
        if len(full_message) <= MAX_MESSAGE_LENGTH:
            await update.effective_message.reply_text(full_message, parse_mode='Markdown')
        else:
            # Dividir a mensagem em partes
            messages = []
            current_message = message

            # Adicionar aulas estudadas
            if study_entries:
                current_message += "*Aulas Estudadas:*\n"
                for entry in study_entries:
                    if len(current_message) + len(entry) + 1 <= MAX_MESSAGE_LENGTH:
                        current_message += entry + "\n"
                    else:
                        messages.append(current_message)
                        current_message = message + "*Aulas Estudadas:*\n" + entry + "\n"
                current_message += "\n"

            # Adicionar revisões
            if review_entries:
                if len(current_message) + len("\n*Revisões Programadas:*\n") <= MAX_MESSAGE_LENGTH:
                    current_message += "\n*Revisões Programadas:*\n"
                else:
                    messages.append(current_message)
                    current_message = message + "\n*Revisões Programadas:*\n"

                for entry in review_entries:
                    if len(current_message) + len(entry) + 1 <= MAX_MESSAGE_LENGTH:
                        current_message += entry + "\n"
                    else:
                        messages.append(current_message)
                        current_message = message + "\n*Revisões Programadas:*\n" + entry + "\n"

            if current_message != message:
                messages.append(current_message)

            # Enviar as mensagens
            for msg in messages:
                await update.effective_message.reply_text(msg, parse_mode='Markdown')
                await asyncio.sleep(0.5)  # Pequeno atraso para evitar flood

    # Se não houver eventos
    if not events_by_date:
        message = "Você ainda não registrou nenhum estudo ou revisão."
        if category_filter:
            message += f" na categoria '{category_filter}'."
        await update.effective_message.reply_text(message, parse_mode='Markdown')

# Início da conversa com /delete
async def delete_study(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # Listar todos os estudos registrados
    cursor.execute('''
        SELECT id, subject, topic, study_date
        FROM studies
        WHERE user_id = ?
    ''', (user_id,))
    studies = cursor.fetchall()

    if not studies:
        await update.effective_message.reply_text("Você ainda não registrou nenhum estudo para excluir.", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END

    # Agrupar estudos por data
    studies_by_date = {}
    for study in studies:
        study_id, subject, topic, study_date = study
        if study_date not in studies_by_date:
            studies_by_date[study_date] = []
        study_date_display = convert_to_display_format(study_date)
        study_entry = f"ID: {study_id} - *{subject} ({topic})* - Estudado em: {study_date_display}"
        studies_by_date[study_date].append(study_entry)

    # Ordenar as datas
    sorted_dates = sorted(studies_by_date.keys())

    # Enviar mensagens separadas por data
    MAX_MESSAGE_LENGTH = 4096  # Limite do Telegram
    for date in sorted_dates:
        date_display = convert_to_display_format(date)
        message = f"*🗑️ Estudos ({date_display}):*\n\n"
        study_entries = studies_by_date[date]
        
        # Dividir as entradas em mensagens menores, se necessário
        current_message = message
        messages = []
        for entry in study_entries:
            if len(current_message) + len(entry) + 1 <= MAX_MESSAGE_LENGTH:
                current_message += entry + "\n"
            else:
                messages.append(current_message)
                current_message = message + entry + "\n"
        if current_message != message:
            messages.append(current_message)

        # Enviar as mensagens
        for msg in messages:
            await update.effective_message.reply_text(msg, reply_markup=ReplyKeyboardRemove(), parse_mode='Markdown')
            await asyncio.sleep(0.5)  # Pequeno atraso para evitar flood

    await update.effective_message.reply_text("Digite o ID do estudo que deseja excluir (ex.: 1):", parse_mode='Markdown')
    return DELETE_STUDY_ID

# Etapa 1: Receber o ID do estudo a ser excluído
async def get_delete_study_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    try:
        study_id = int(update.message.text)
        # Verificar se o ID existe
        cursor.execute('''
            SELECT id, subject, topic, study_date
            FROM studies
            WHERE user_id = ? AND id = ?
        ''', (user_id, study_id))
        study = cursor.fetchone()

        if not study:
            await update.message.reply_text("ID inválido. Por favor, escolha um ID válido da lista.")
            return DELETE_STUDY_ID

        context.user_data['delete_study_id'] = study_id
        study_date_display = convert_to_display_format(study[3])
        await update.message.reply_text(
            f"Você está prestes a excluir o estudo *{study[1]} ({study[2]})* - Estudado em: {study_date_display}.\n"
            "Isso também excluirá todas as revisões associadas. Confirma a exclusão? (sim/não)",
            parse_mode='Markdown'
        )
        return DELETE_CONFIRM
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido para o ID (ex.: 1).")
        return DELETE_STUDY_ID

# Etapa 2: Confirmar a exclusão
async def get_delete_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    response = update.message.text.lower()
    if response not in ['sim', 'não', 'nao']:
        await update.message.reply_text("Por favor, responda 'sim' ou 'não'.")
        return DELETE_CONFIRM

    if response == 'sim':
        study_id = context.user_data['delete_study_id']
        # Excluir as revisões associadas
        cursor.execute('DELETE FROM reviews WHERE study_id = ?', (study_id,))
        # Excluir o estudo
        cursor.execute('DELETE FROM studies WHERE id = ?', (study_id,))
        conn.commit()
        await update.message.reply_text("Estudo e revisões associadas excluídos com sucesso!", parse_mode='Markdown')
    else:
        await update.message.reply_text("Exclusão cancelada.", parse_mode='Markdown')

    context.user_data.clear()
    return ConversationHandler.END

async def stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # Total de questões feitas
    cursor.execute('SELECT SUM(questions_studied) FROM studies WHERE user_id = ?', (user_id,))
    total_questions = cursor.fetchone()[0] or 0

    # Número de revisões concluídas
    cursor.execute('SELECT COUNT(*) FROM reviews WHERE completed = 1 AND study_id IN (SELECT id FROM studies WHERE user_id = ?)', (user_id,))
    completed_reviews = cursor.fetchone()[0]

    # Matéria mais estudada
    cursor.execute('''
        SELECT subject, COUNT(*) as count
        FROM studies
        WHERE user_id = ?
        GROUP BY subject
        ORDER BY count DESC
        LIMIT 1
    ''', (user_id,))
    most_studied = cursor.fetchone()
    most_studied_subject = most_studied[0] if most_studied else "Nenhuma"
    most_studied_count = most_studied[1] if most_studied else 0

    message = "*📊 Estatísticas de Estudo*\n\n"
    message += f"- **Total de questões feitas:** {total_questions}\n"
    message += f"- **Revisões concluídas:** {completed_reviews}\n"
    message += f"- **Matéria mais estudada:** {most_studied_subject} ({most_studied_count} registros)\n"

    # Dividir a mensagem, se necessário (embora atualmente seja curta)
    MAX_MESSAGE_LENGTH = 4096  # Limite do Telegram
    if len(message) <= MAX_MESSAGE_LENGTH:
        await update.effective_message.reply_text(message, parse_mode='Markdown')
    else:
        parts = []
        current_part = ""
        for line in message.split('\n'):
            if len(current_part) + len(line) + 1 <= MAX_MESSAGE_LENGTH:
                current_part += line + '\n'
            else:
                parts.append(current_part)
                current_part = line + '\n'
        if current_part:
            parts.append(current_part)

        for part in parts:
            await update.effective_message.reply_text(part, parse_mode='Markdown')
            await asyncio.sleep(0.5)  # Pequeno atraso para evitar flood

# Início da conversa com /markdone
async def mark_done(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # Listar todas as revisões não concluídas
    cursor.execute('''
        SELECT s.subject, s.topic, r.id, r.review_number, r.review_date
        FROM studies s
        JOIN reviews r ON s.id = r.study_id
        WHERE s.user_id = ? AND r.completed = 0
    ''', (user_id,))
    reviews = cursor.fetchall()

    if not reviews:
        await update.effective_message.reply_text("Você não tem revisões pendentes para marcar como concluídas.", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END

    # Agrupar revisões por data
    reviews_by_date = {}
    for review in reviews:
        subject, topic, review_id, review_number, review_date = review
        if review_date not in reviews_by_date:
            reviews_by_date[review_date] = []
        review_date_display = convert_to_display_format(review_date)
        review_entry = f"ID: {review_id} - *{subject} ({topic})* - Revisão {review_number} - Data: {review_date_display}"
        reviews_by_date[review_date].append(review_entry)

    # Ordenar as datas
    sorted_dates = sorted(reviews_by_date.keys())

    # Enviar mensagens separadas por data
    MAX_MESSAGE_LENGTH = 4096  # Limite do Telegram
    for date in sorted_dates:
        date_display = convert_to_display_format(date)
        message = f"*✅ Revisões Pendentes ({date_display}):*\n\n"
        message += "\n".join(reviews_by_date[date]) + "\n"

        # Dividir a mensagem em partes menores, se necessário
        if len(message) <= MAX_MESSAGE_LENGTH:
            await update.effective_message.reply_text(message, reply_markup=ReplyKeyboardRemove(), parse_mode='Markdown')
        else:
            parts = []
            current_part = ""
            for line in message.split('\n'):
                if len(current_part) + len(line) + 1 <= MAX_MESSAGE_LENGTH:
                    current_part += line + '\n'
                else:
                    parts.append(current_part)
                    current_part = line + '\n'
            if current_part:
                parts.append(current_part)

            for part in parts:
                await update.effective_message.reply_text(part, reply_markup=ReplyKeyboardRemove(), parse_mode='Markdown')
                await asyncio.sleep(0.5)  # Pequeno atraso para evitar flood

    await update.effective_message.reply_text("Digite o ID da revisão que deseja marcar como concluída (ex.: 1):", parse_mode='Markdown')
    return MARKDONE_REVIEW_ID

# Etapa 1: Receber o ID da revisão a ser marcada como concluída
async def get_markdone_review_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    try:
        review_id = int(update.message.text)
        # Verificar se o ID existe
        cursor.execute('''
            SELECT s.subject, s.topic, r.id, r.review_number, r.review_date
            FROM studies s
            JOIN reviews r ON s.id = r.study_id
            WHERE s.user_id = ? AND r.id = ? AND r.completed = 0
        ''', (user_id, review_id))
        review = cursor.fetchone()

        if not review:
            await update.message.reply_text("ID inválido ou revisão já concluída. Por favor, escolha um ID válido da lista.")
            return MARKDONE_REVIEW_ID

        context.user_data['markdone_review_id'] = review_id
        review_date_display = convert_to_display_format(review[4])
        await update.message.reply_text(
            f"Você está prestes a marcar como concluída a revisão *{review[0]} ({review[1]})* - Revisão {review[3]} - Data: {review_date_display}.\n"
            "Confirma? (sim/não)",
            parse_mode='Markdown'
        )
        return MARKDONE_CONFIRM
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido para o ID (ex.: 1).")
        return MARKDONE_REVIEW_ID

# Etapa 2: Confirmar a marcação como concluída
async def get_markdone_confirm(update: Update, context: ContextTypes.DEFAULT_TYPE):
    response = update.message.text.lower()
    if response not in ['sim', 'não', 'nao']:
        await update.message.reply_text("Por favor, responda 'sim' ou 'não'.")
        return MARKDONE_CONFIRM

    if response == 'sim':
        review_id = context.user_data['markdone_review_id']
        cursor.execute('UPDATE reviews SET completed = 1 WHERE id = ?', (review_id,))
        conn.commit()
        await update.message.reply_text("Revisão marcada como concluída com sucesso! ✅", parse_mode='Markdown')
    else:
        await update.message.reply_text("Marcação cancelada.", parse_mode='Markdown')

    context.user_data.clear()
    return ConversationHandler.END

async def weekly_report(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id
    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday())  # Início da semana (segunda-feira)
    end_of_week = start_of_week + timedelta(days=6)  # Fim da semana (domingo)

    start_date = start_of_week.strftime('%Y-%m-%d')
    end_date = end_of_week.strftime('%Y-%m-%d')
    start_date_display = convert_to_display_format(start_date)
    end_date_display = convert_to_display_format(end_date)

    # Estudos realizados na semana
    cursor.execute('''
        SELECT subject, topic, study_date
        FROM studies
        WHERE user_id = ? AND study_date BETWEEN ? AND ?
    ''', (user_id, start_date, end_date))
    studies = cursor.fetchall()

    # Revisões concluídas na semana
    cursor.execute('''
        SELECT s.subject, s.topic, r.review_date
        FROM studies s
        JOIN reviews r ON s.id = r.study_id
        WHERE s.user_id = ? AND r.review_date BETWEEN ? AND ? AND r.completed = 1
    ''', (user_id, start_date, end_date))
    completed_reviews = cursor.fetchall()

    # Matéria mais estudada na semana
    cursor.execute('''
        SELECT subject, COUNT(*) as count
        FROM studies
        WHERE user_id = ? AND study_date BETWEEN ? AND ?
        GROUP BY subject
        ORDER BY count DESC
        LIMIT 1
    ''', (user_id, start_date, end_date))
    most_studied = cursor.fetchone()
    most_studied_subject = most_studied[0] if most_studied else "Nenhuma"
    most_studied_count = most_studied[1] if most_studied else 0

    message = f"*📅 Relatório Semanal ({start_date_display} a {end_date_display})*\n\n"
    message += f"- **Estudos realizados:** {len(studies)}\n"

    # Preparar entradas de estudos
    study_entries = []
    for study in studies:
        subject, topic, study_date = study
        study_date_display = convert_to_display_format(study_date)
        entry = f"  - *{subject} ({topic})* - Estudado em: {study_date_display}"
        study_entries.append(entry)

    # Preparar entradas de revisões concluídas
    review_entries = []
    for review in completed_reviews:
        subject, topic, review_date = review
        review_date_display = convert_to_display_format(review_date)
        entry = f"  - *{subject} ({topic})* - Revisão concluída em: {review_date_display}"
        review_entries.append(entry)

    # Adicionar resumo de revisões concluídas
    message += f"- **Revisões concluídas:** {len(completed_reviews)}\n"

    # Adicionar matéria mais estudada
    message += f"- **Matéria mais estudada:** {most_studied_subject} ({most_studied_count} registros)\n"

    # Dividir a mensagem em partes
    MAX_MESSAGE_LENGTH = 4096  # Limite do Telegram
    messages = []
    current_message = message

    # Adicionar estudos
    if study_entries:
        current_message += "\n*Estudos Realizados:*\n"
        for entry in study_entries:
            if len(current_message) + len(entry) + 1 <= MAX_MESSAGE_LENGTH:
                current_message += entry + "\n"
            else:
                messages.append(current_message)
                current_message = f"*📅 Relatório Semanal ({start_date_display} a {end_date_display}) - Continuação*\n\n*Estudos Realizados:*\n" + entry + "\n"
        current_message += "\n"

    # Adicionar revisões concluídas
    if review_entries:
        if len(current_message) + len("\n*Revisões Concluídas:*\n") <= MAX_MESSAGE_LENGTH:
            current_message += "\n*Revisões Concluídas:*\n"
        else:
            messages.append(current_message)
            current_message = f"*📅 Relatório Semanal ({start_date_display} a {end_date_display}) - Continuação*\n\n*Revisões Concluídas:*\n"

        for entry in review_entries:
            if len(current_message) + len(entry) + 1 <= MAX_MESSAGE_LENGTH:
                current_message += entry + "\n"
            else:
                messages.append(current_message)
                current_message = f"*📅 Relatório Semanal ({start_date_display} a {end_date_display}) - Continuação*\n\n*Revisões Concluídas:*\n" + entry + "\n"

    if current_message != message:
        messages.append(current_message)

    # Enviar as mensagens
    for msg in messages:
        await update.effective_message.reply_text(msg, parse_mode='Markdown')
        await asyncio.sleep(0.5)  # Pequeno atraso para evitar flood
    user_id = update.effective_user.id

    # Listar todos os estudos registrados
    cursor.execute('''
        SELECT id, subject, topic, study_date
        FROM studies
        WHERE user_id = ?
    ''', (user_id,))
    studies = cursor.fetchall()

    if not studies:
        await update.effective_message.reply_text("Você ainda não registrou nenhum estudo para editar.", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END

    message = "*📝 Selecione o estudo que deseja editar:*\n\n"
    for study in studies:
        study_id, subject, topic, study_date = study
        study_date_display = convert_to_display_format(study_date)
        message += f"ID: {study_id} - *{subject} ({topic})* - Estudado em: {study_date_display}\n"

    message += "\nDigite o ID do estudo que deseja editar (ex.: 1):"
    await update.effective_message.reply_text(message, reply_markup=ReplyKeyboardRemove(), parse_mode='Markdown')
    return EDIT_STUDY_ID

# Etapa 1: Receber o ID do estudo a ser editado
async def get_study_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    try:
        study_id = int(update.message.text)
        # Verificar se o ID existe
        cursor.execute('''
            SELECT id, subject, topic, study_date
            FROM studies
            WHERE user_id = ? AND id = ?
        ''', (user_id, study_id))
        study = cursor.fetchone()

        if not study:
            await update.message.reply_text("ID inválido. Por favor, escolha um ID válido da lista.")
            return EDIT_STUDY_ID

        context.user_data['edit_study_id'] = study_id
        context.user_data['edit_study_date'] = study[3]  # Salvar a data de estudo original
        await update.message.reply_text("Você deseja editar a data de estudo? Responda 'sim' ou 'não'.")
        return EDIT_DATE_YN
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido para o ID (ex.: 1).")
        return EDIT_STUDY_ID

# Etapa 2: Perguntar se deseja editar a data de estudo
async def get_edit_date_yn(update: Update, context: ContextTypes.DEFAULT_TYPE):
    response = update.message.text.lower()
    if response not in ['sim', 'não', 'nao']:
        await update.message.reply_text("Por favor, responda 'sim' ou 'não'.")
        return EDIT_DATE_YN
    context.user_data['edit_date_yn'] = response
    if response == 'sim':
        await update.message.reply_text("Digite a nova data de estudo no formato DD-MM-YYYY (ex.: 15-04-2025):")
        return EDIT_DATE
    else:
        await update.message.reply_text("Quantas revisões você deseja programar para este estudo? (Ex.: 3)")
        return EDIT_NUM_REVIEWS

# Etapa 3: Receber a nova data de estudo e recalcular as revisões
async def get_edit_date(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        new_date = update.message.text
        # Validar o formato da data e converter para o formato interno
        new_date_db = convert_to_db_format(new_date)
        old_date = context.user_data['edit_study_date']
        study_id = context.user_data['edit_study_id']

        # Buscar todas as revisões do estudo
        cursor.execute('''
            SELECT review_date, review_questions, review_number
            FROM reviews
            WHERE study_id = ?
        ''', (study_id,))
        reviews = cursor.fetchall()

        # Calcular os intervalos de dias entre a data de estudo original e cada revisão
        intervals = []
        for review in reviews:
            review_date, review_questions, review_number = review
            days = calculate_days_difference(old_date, review_date)
            intervals.append((days, review_questions, review_number))

        # Atualizar a data de estudo no banco de dados
        cursor.execute('''
            UPDATE studies
            SET study_date = ?
            WHERE id = ?
        ''', (new_date_db, study_id))

        # Recalcular as datas das revisões com base na nova data de estudo
        cursor.execute('DELETE FROM reviews WHERE study_id = ?', (study_id,))
        for days, questions, review_number in intervals:
            new_review_date = calculate_review_date(new_date_db, days)
            cursor.execute('''
                INSERT INTO reviews (study_id, review_number, review_date, review_questions)
                VALUES (?, ?, ?, ?)
            ''', (study_id, review_number, new_review_date, questions))

        conn.commit()

        context.user_data['edit_study_date'] = new_date_db
        new_date_display = convert_to_display_format(new_date_db)
        await update.message.reply_text(f"Data de estudo atualizada para {new_date_display}! As datas das revisões foram recalculadas.")
        await update.message.reply_text("Quantas revisões você deseja programar para este estudo? (Ex.: 3)")
        return EDIT_NUM_REVIEWS
    except ValueError:
        await update.message.reply_text("Por favor, digite uma data válida no formato DD-MM-YYYY (ex.: 15-04-2025).")
        return EDIT_DATE

# Etapa 4: Receber o número de revisões
async def get_edit_num_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        num_reviews = int(update.message.text)
        if num_reviews < 1:
            raise ValueError
        context.user_data['edit_num_reviews'] = num_reviews
        context.user_data['edit_reviews'] = []
        context.user_data['current_review'] = 1
        await update.message.reply_text(f"Em quantos dias você quer fazer a {context.user_data['current_review']}ª revisão? (Ex.: 3)")
        return EDIT_REVIEW_DAYS
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido maior que 0 (ex.: 3).")
        return EDIT_NUM_REVIEWS

# Etapa 5: Receber os dias para a revisão atual
async def get_edit_review_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        days = int(update.message.text)
        if days < 0:
            raise ValueError
        context.user_data['edit_reviews'].append({'days': days})
        await update.message.reply_text(f"Quantas questões você quer responder na {context.user_data['current_review']}ª revisão? (Ex.: 15)")
        return EDIT_REVIEW_QUESTIONS
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido (ex.: 3).")
        return EDIT_REVIEW_DAYS

# Etapa 6: Receber as questões para a revisão atual
async def get_edit_review_questions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        questions = int(update.message.text)
        if questions < 0:
            raise ValueError
        context.user_data['edit_reviews'][-1]['questions'] = questions
        await update.message.reply_text(f"Qual a prioridade desta revisão? (alta, média, baixa. Digite 'média' para padrão.)")
        return EDIT_REVIEW_PRIORITY
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido (ex.: 15).")
        return EDIT_REVIEW_QUESTIONS

# Etapa 7: Receber a prioridade da revisão e continuar ou finalizar
async def get_edit_review_priority(update: Update, context: ContextTypes.DEFAULT_TYPE):
    priority = update.message.text.lower()
    if priority not in ['alta', 'média', 'baixa']:
        await update.message.reply_text("Por favor, escolha uma prioridade válida: alta, média ou baixa.")
        return EDIT_REVIEW_PRIORITY
    context.user_data['edit_reviews'][-1]['priority'] = priority

    # Incrementar o contador de revisões
    context.user_data['current_review'] += 1

    # Verificar se ainda há revisões a serem adicionadas
    if context.user_data['current_review'] <= context.user_data['edit_num_reviews']:
        await update.message.reply_text(f"Em quantos dias você quer fazer a {context.user_data['current_review']}ª revisão? (Ex.: 3)")
        return EDIT_REVIEW_DAYS
    else:
        # Remover as revisões antigas do estudo
        study_id = context.user_data['edit_study_id']
        cursor.execute('DELETE FROM reviews WHERE study_id = ?', (study_id,))

        # Adicionar as novas revisões
        message = "*Revisões atualizadas com sucesso!* Novas datas:\n"
        for i, review in enumerate(context.user_data['edit_reviews'], 1):
            review_date = calculate_review_date(context.user_data['edit_study_date'], review['days'])
            review_date_display = convert_to_display_format(review_date)
            cursor.execute('''
                INSERT INTO reviews (study_id, review_number, review_date, review_questions, priority)
                VALUES (?, ?, ?, ?, ?)
            ''', (study_id, i, review_date, review['questions'], review['priority']))
            message += f"Revisão {i}: {review_date_display} ({review['questions']} questões, Prioridade: {review['priority']})\n"

        conn.commit()

        await update.message.reply_text(message, parse_mode='Markdown')

        # Limpar os dados temporários
        context.user_data.clear()
        return ConversationHandler.END

async def redo_study(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # Listar todos os estudos registrados
    cursor.execute('''
        SELECT id, subject, topic, study_date
        FROM studies
        WHERE user_id = ?
    ''', (user_id,))
    studies = cursor.fetchall()

    if not studies:
        await update.effective_message.reply_text("Você ainda não registrou nenhum estudo para refazer o plano.", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END

    message = "*🔄 Selecione o estudo para refazer o plano de revisões:*\n\n"
    for study in studies:
        study_id, subject, topic, study_date = study
        study_date_display = convert_to_display_format(study_date)
        message += f"ID: {study_id} - *{subject} ({topic})* - Estudado em: {study_date_display}\n"

    message += "\nDigite o ID do estudo que deseja refazer o plano (ex.: 1):"
    await update.effective_message.reply_text(message, reply_markup=ReplyKeyboardRemove(), parse_mode='Markdown')
    return REDO_STUDY_ID

# Etapa 1: Receber o ID do estudo para refazer o plano
async def get_redo_study_id(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.message.from_user.id
    try:
        study_id = int(update.message.text)
        # Verificar se o ID existe
        cursor.execute('''
            SELECT id, subject, topic, study_date
            FROM studies
            WHERE user_id = ? AND id = ?
        ''', (user_id, study_id))
        study = cursor.fetchone()

        if not study:
            await update.message.reply_text("ID inválido. Por favor, escolha um ID válido da lista.")
            return REDO_STUDY_ID

        context.user_data['redo_study_id'] = study_id
        context.user_data['redo_study_date'] = study[3]  # Salvar a data de estudo para recalcular as revisões
        await update.message.reply_text("Quantas revisões você deseja programar para este estudo? (Ex.: 3)")
        return REDO_NUM_REVIEWS
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido para o ID (ex.: 1).")
        return REDO_STUDY_ID

# Etapa 2: Receber o número de revisões para refazer o plano
async def get_redo_num_reviews(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        num_reviews = int(update.message.text)
        if num_reviews < 1:
            raise ValueError
        context.user_data['redo_num_reviews'] = num_reviews
        context.user_data['redo_reviews'] = []
        context.user_data['current_review'] = 1
        await update.message.reply_text(f"Em quantos dias você quer fazer a {context.user_data['current_review']}ª revisão? (Ex.: 3)")
        return REDO_REVIEW_DAYS
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido maior que 0 (ex.: 3).")
        return REDO_NUM_REVIEWS

# Etapa 3: Receber os dias para a revisão atual (refazer o plano)
async def get_redo_review_days(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        days = int(update.message.text)
        if days < 0:
            raise ValueError
        context.user_data['redo_reviews'].append({'days': days})
        await update.message.reply_text(f"Quantas questões você quer responder na {context.user_data['current_review']}ª revisão? (Ex.: 15)")
        return REDO_REVIEW_QUESTIONS
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido (ex.: 3).")
        return REDO_REVIEW_DAYS

# Etapa 4: Receber as questões para a revisão atual
async def get_redo_review_questions(update: Update, context: ContextTypes.DEFAULT_TYPE):
    try:
        questions = int(update.message.text)
        if questions < 0:
            raise ValueError
        context.user_data['redo_reviews'][-1]['questions'] = questions
        await update.message.reply_text(f"Qual a prioridade desta revisão? (alta, média, baixa. Digite 'média' para padrão.)")
        return REDO_REVIEW_PRIORITY
    except ValueError:
        await update.message.reply_text("Por favor, digite um número válido (ex.: 15).")
        return REDO_REVIEW_QUESTIONS

# Etapa 5: Receber a prioridade da revisão e continuar ou finalizar
async def get_redo_review_priority(update: Update, context: ContextTypes.DEFAULT_TYPE):
    priority = update.message.text.lower()
    if priority not in ['alta', 'média', 'baixa']:
        await update.message.reply_text("Por favor, escolha uma prioridade válida: alta, média ou baixa.")
        return REDO_REVIEW_PRIORITY
    context.user_data['redo_reviews'][-1]['priority'] = priority

    # Incrementar o contador de revisões
    context.user_data['current_review'] += 1

    # Verificar se ainda há revisões a serem adicionadas
    if context.user_data['current_review'] <= context.user_data['redo_num_reviews']:
        await update.message.reply_text(f"Em quantos dias você quer fazer a {context.user_data['current_review']}ª revisão? (Ex.: 3)")
        return REDO_REVIEW_DAYS
    else:
        # Remover as revisões antigas do estudo
        study_id = context.user_data['redo_study_id']
        cursor.execute('DELETE FROM reviews WHERE study_id = ?', (study_id,))

        # Adicionar as novas revisões
        message = "*Plano de revisões refeito com sucesso!* Novas datas:\n"
        for i, review in enumerate(context.user_data['redo_reviews'], 1):
            review_date = calculate_review_date(context.user_data['redo_study_date'], review['days'])
            review_date_display = convert_to_display_format(review_date)
            cursor.execute('''
                INSERT INTO reviews (study_id, review_number, review_date, review_questions, priority)
                VALUES (?, ?, ?, ?, ?)
            ''', (study_id, i, review_date, review['questions'], review['priority']))
            message += f"Revisão {i}: {review_date_display} ({review['questions']} questões, Prioridade: {review['priority']})\n"

        conn.commit()

        await update.message.reply_text(message, parse_mode='Markdown')

        # Limpar os dados temporários
        context.user_data.clear()
        return ConversationHandler.END
# Início da conversa com /edit
async def edit_study(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # Listar todos os estudos registrados
    cursor.execute('''
        SELECT id, subject, topic, study_date
        FROM studies
        WHERE user_id = ?
    ''', (user_id,))
    studies = cursor.fetchall()

    if not studies:
        await update.effective_message.reply_text("Você ainda não registrou nenhum estudo para editar.", reply_markup=ReplyKeyboardRemove())
        return ConversationHandler.END

    message = "*📝 Selecione o estudo que deseja editar:*\n\n"
    for study in studies:
        study_id, subject, topic, study_date = study
        study_date_display = convert_to_display_format(study_date)
        message += f"ID: {study_id} - *{subject} ({topic})* - Estudado em: {study_date_display}\n"

    message += "\nDigite o ID do estudo que deseja editar (ex.: 1):"
    await update.effective_message.reply_text(message, reply_markup=ReplyKeyboardRemove(), parse_mode='Markdown')
    return EDIT_STUDY_ID

# Comando /export para exportar dados para Excel
async def export_data(update: Update, context: ContextTypes.DEFAULT_TYPE):
    user_id = update.effective_user.id

    # Buscar todos os estudos
    cursor.execute('''
        SELECT id, user_id, subject, topic, study_date, questions_studied, category, notes
        FROM studies
        WHERE user_id = ?
    ''', (user_id,))
    studies = cursor.fetchall()

    # Buscar todas as revisões
    cursor.execute('''
        SELECT s.subject, s.topic, r.id, r.study_id, r.review_number, r.review_date, r.review_questions, r.priority, r.notes, r.completed
        FROM studies s
        JOIN reviews r ON s.id = r.study_id
        WHERE s.user_id = ?
    ''', (user_id,))
    reviews = cursor.fetchall()

    # Criar um novo arquivo Excel
    workbook = openpyxl.Workbook()

    # Remover a aba padrão criada automaticamente
    default_sheet = workbook.active
    workbook.remove(default_sheet)

    # Criar aba para Estudos
    sheet_studies = workbook.create_sheet(title="Estudos")

    # Adicionar título do relatório na aba Estudos
    export_date = datetime.now().strftime('%d-%m-%Y %H:%M:%S')
    sheet_studies['A1'] = f"Relatório de Estudos - Exportado em {export_date}"
    sheet_studies['A1'].font = Font(bold=True, size=14)
    sheet_studies.merge_cells('A1:I1')
    sheet_studies['A1'].alignment = Alignment(horizontal='center')

    # Adicionar seção de Estudos
    row = 3
    sheet_studies[f'A{row}'] = "Lista de Estudos"
    sheet_studies[f'A{row}'].font = Font(bold=True, size=12)
    sheet_studies.merge_cells(f'A{row}:I{row}')
    row += 1

    # Cabeçalhos para Estudos
    headers_studies = [
        "Índice", "ID", "Usuário", "Disciplina", "Tema", "Data de Estudo", "Questões Feitas", "Categoria", "Notas"
    ]
    for col, header in enumerate(headers_studies, 1):
        cell = sheet_studies.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Adicionar dados dos Estudos
    row += 1
    for idx, study in enumerate(studies, 1):
        study_date_display = convert_to_display_format(study[4])
        notes = study[7].replace('\n', ' ') if study[7] else ''  # Remover quebras de linha
        data = [
            idx, study[0], study[1], study[2], study[3], study_date_display, study[5], study[6] or '', notes
        ]
        for col, value in enumerate(data, 1):
            sheet_studies.cell(row=row, column=col).value = value
        row += 1

    # Ajustar largura das colunas para Estudos
    column_widths = {}
    for row in sheet_studies.rows:
        for cell in row:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Ignorar células mescladas
                column = cell.column  # Usar o índice da coluna
                column_letter = openpyxl.utils.get_column_letter(column)
                try:
                    length = len(str(cell.value))
                    column_widths[column_letter] = max(column_widths.get(column_letter, 0), length)
                except:
                    pass

    for column_letter, width in column_widths.items():
        adjusted_width = (width + 2)
        sheet_studies.column_dimensions[column_letter].width = adjusted_width

    # Adicionar filtros para Estudos
    sheet_studies.auto_filter.ref = f"A4:I{4 + len(studies)}"

    # Criar aba para Revisões
    sheet_reviews = workbook.create_sheet(title="Revisões")

    # Adicionar título do relatório na aba Revisões
    sheet_reviews['A1'] = f"Relatório de Revisões - Exportado em {export_date}"
    sheet_reviews['A1'].font = Font(bold=True, size=14)
    sheet_reviews.merge_cells('A1:K1')
    sheet_reviews['A1'].alignment = Alignment(horizontal='center')

    # Adicionar seção de Revisões
    row = 3
    sheet_reviews[f'A{row}'] = "Lista de Revisões"
    sheet_reviews[f'A{row}'].font = Font(bold=True, size=12)
    sheet_reviews.merge_cells(f'A{row}:K{row}')
    row += 1

    # Cabeçalhos para Revisões
    headers_reviews = [
        "Índice", "ID", "Disciplina", "Tema", "ID do Estudo", "Número da Revisão", "Data da Revisão", "Questões", "Prioridade", "Notas", "Status"
    ]
    for col, header in enumerate(headers_reviews, 1):
        cell = sheet_reviews.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal='center')

    # Adicionar dados das Revisões
    row += 1
    for idx, review in enumerate(reviews, 1):
        review_date_display = convert_to_display_format(review[5])
        status = "Concluída" if review[9] else "Pendente"
        notes = review[8].replace('\n', ' ') if review[8] else ''  # Remover quebras de linha
        data = [
            idx, review[2], review[0], review[1], review[3], review[4], review_date_display, review[6], review[7], notes, status
        ]
        for col, value in enumerate(data, 1):
            sheet_reviews.cell(row=row, column=col).value = value
        row += 1

    # Ajustar largura das colunas para Revisões
    column_widths = {}
    for row in sheet_reviews.rows:
        for cell in row:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Ignorar células mescladas
                column = cell.column  # Usar o índice da coluna
                column_letter = openpyxl.utils.get_column_letter(column)
                try:
                    length = len(str(cell.value))
                    column_widths[column_letter] = max(column_widths.get(column_letter, 0), length)
                except:
                    pass

    for column_letter, width in column_widths.items():
        adjusted_width = (width + 2)
        sheet_reviews.column_dimensions[column_letter].width = adjusted_width

    # Adicionar filtros para Revisões
    sheet_reviews.auto_filter.ref = f"A4:K{4 + len(reviews)}"

    # Criar aba para Resumo
    sheet_summary = workbook.create_sheet(title="Resumo")

    # Adicionar título do relatório na aba Resumo
    sheet_summary['A1'] = f"Resumo - Exportado em {export_date}"
    sheet_summary['A1'].font = Font(bold=True, size=14)
    sheet_summary.merge_cells('A1:C1')
    sheet_summary['A1'].alignment = Alignment(horizontal='center')

    # Adicionar seção de Resumo
    row = 3
    sheet_summary[f'A{row}'] = "Resumo"
    sheet_summary[f'A{row}'].font = Font(bold=True, size=12)
    sheet_summary.merge_cells(f'A{row}:C{row}')
    row += 1

    total_studies = len(studies)
    total_reviews = len(reviews)
    total_completed_reviews = sum(1 for review in reviews if review[9])
    summary_data = [
        ["Total de Estudos", total_studies, ""],
        ["Total de Revisões", total_reviews, ""],
        ["Revisões Concluídas", total_completed_reviews, ""]
    ]
    for summary_row in summary_data:
        for col, value in enumerate(summary_row, 1):
            cell = sheet_summary.cell(row=row, column=col)
            cell.value = value
            if col == 1:
                cell.font = Font(bold=True)
        row += 1

    # Ajustar largura das colunas para Resumo
    column_widths = {}
    for row in sheet_summary.rows:
        for cell in row:
            if cell.value and not isinstance(cell, openpyxl.cell.cell.MergedCell):  # Ignorar células mescladas
                column = cell.column  # Usar o índice da coluna
                column_letter = openpyxl.utils.get_column_letter(column)
                try:
                    length = len(str(cell.value))
                    column_widths[column_letter] = max(column_widths.get(column_letter, 0), length)
                except:
                    pass

    for column_letter, width in column_widths.items():
        adjusted_width = (width + 2)
        sheet_summary.column_dimensions[column_letter].width = adjusted_width

    # Salvar o arquivo Excel
    excel_filename = f"study_export_{user_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    workbook.save(excel_filename)

    # Enviar o arquivo Excel
    with open(excel_filename, 'rb') as file:
        await update.effective_message.reply_document(document=file, filename=excel_filename, caption="Aqui está o export dos seus estudos e revisões!")

    # Remover o arquivo temporário
    os.remove(excel_filename)

# Função principal para rodar o bot
def main():
    # Criar a aplicação com timeout aumentado
    app = Application.builder().token(TOKEN).read_timeout(30).write_timeout(30).connect_timeout(30).build()
    print("[Main] Aplicação criada com sucesso.")

    # Configurar a conversa para o comando /add
    add_handler = ConversationHandler(
        entry_points=[CommandHandler('add', add_study)],
        states={
            SUBJECT: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_subject)],
            TOPIC: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_topic)],
            CATEGORY: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_category)],
            QUESTIONS_YN: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_questions_yn)],
            QUESTIONS_AMOUNT: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_questions_amount)],
            NUM_REVIEWS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_num_reviews)],
            REVIEW_DAYS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_review_days)],
            REVIEW_QUESTIONS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_review_questions)],
            REVIEW_PRIORITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_review_priority)],
            REVIEW_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_review_notes)],
            STUDY_NOTES: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_study_notes)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )

    # Configurar a conversa para o comando /edit
    edit_handler = ConversationHandler(
        entry_points=[CommandHandler('edit', edit_study)],
        states={
            EDIT_STUDY_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_study_id)],
            EDIT_DATE_YN: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_edit_date_yn)],
            EDIT_DATE: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_edit_date)],
            EDIT_NUM_REVIEWS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_edit_num_reviews)],
            EDIT_REVIEW_DAYS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_edit_review_days)],
            EDIT_REVIEW_QUESTIONS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_edit_review_questions)],
            EDIT_REVIEW_PRIORITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_edit_review_priority)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )

    # Configurar a conversa para o comando /redo
    redo_handler = ConversationHandler(
        entry_points=[CommandHandler('redo', redo_study)],
        states={
            REDO_STUDY_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_redo_study_id)],
            REDO_NUM_REVIEWS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_redo_num_reviews)],
            REDO_REVIEW_DAYS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_redo_review_days)],
            REDO_REVIEW_QUESTIONS: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_redo_review_questions)],
            REDO_REVIEW_PRIORITY: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_redo_review_priority)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )

    # Configurar a conversa para o comando /delete
    delete_handler = ConversationHandler(
        entry_points=[CommandHandler('delete', delete_study)],
        states={
            DELETE_STUDY_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_delete_study_id)],
            DELETE_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_delete_confirm)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )

    # Configurar a conversa para o comando /markdone
    markdone_handler = ConversationHandler(
        entry_points=[CommandHandler('markdone', mark_done)],
        states={
            MARKDONE_REVIEW_ID: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_markdone_review_id)],
            MARKDONE_CONFIRM: [MessageHandler(filters.TEXT & ~filters.COMMAND, get_markdone_confirm)],
        },
        fallbacks=[CommandHandler('cancel', cancel)]
    )

    # Adicionar comandos
    app.add_handler(CommandHandler("start", start))
    app.add_handler(CommandHandler("stop", stop))
    app.add_handler(add_handler)
    app.add_handler(edit_handler)
    app.add_handler(redo_handler)
    app.add_handler(delete_handler)
    app.add_handler(markdone_handler)
    app.add_handler(CommandHandler("check", check_reviews))
    app.add_handler(CommandHandler("summary", summary))
    app.add_handler(CommandHandler("stats", stats))
    app.add_handler(CommandHandler("export", export_data))
    app.add_handler(CommandHandler("weeklyreport", weekly_report))

    # Iniciar o bot
    print("[Main] Bot está rodando...")
    app.run_polling()
    print("[Main] Bot parou de rodarr.")

if __name__ == '__main__':
    main()